import time
from datetime import datetime, date

import pandas as pd
import win32com.client
from openpyxl import load_workbook
from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait


def send_report(table, office):
    report = '''<!DOCTYPE html>
    <html lang="en">
    <head>
      <meta charset="UTF-8">
      <title>Bond Manifest Report</title>
    </head>
    <body>
    <style>
table {
    border-collapse: collapse;
    width: 100%;
}
th, td {
    text-align: left;
    padding: 8px;
}
tr:nth-child(even) {background-color: #f2f2f2;}


</style>
    <h2>Bond Manifest report</h2>
    <table class="table table-stripped table-responsive table-bordered" border="1">
      <tr>
        <th>TN</th>
        <th>MAWB</th>
        <th>HAWB</th>
        <th>Cnee Code</th>
        <th>ETA</th>
        <th>Dest.</th>
        <th>Goods at</th>
        <th>Status</th>
      </tr>
      '''
    report_end = '</table></body></html>'
    report = report + table + report_end
    outlook_con = win32com.client.Dispatch('outlook.application')
    rules = load_workbook('rules.xlsx', read_only=True)
    contacts = rules['Contacts']

    mail = outlook_con.CreateItem(0)
    for line in range(2, 20):
        if contacts['A' + str(line)].value == office:
            mail.To = contacts['B' + str(line)].value
    branch = office[19:24]
    mail.CC = 'aleksei.shcherbakov@kuehne-nagel.com'
    mail.Subject = branch + ' Bond Manifest report for ' + str(date.today())
    mail.HTMLBody = report
    mail.Send()
    print('Email Generated')


def open_bond_page(tn, branch, from_port, to_port, goods_at, eta, eta_warehouse):
    global ccn_change
    options = Options()
    options.add_argument("--headless")
    driver = webdriver.Firefox(firefox_options=options)
    timestamp = int(time.time() * 1000)
    url = "http://amer.int.kn/customs/logon.do?dispatch=sso&requestedUrl=%2FbondAdvice.do%3FknDesktop%3DY%26branchCode%3D" + branch + "%26userId%3Daleksei.shcherbakov%26departmentCode%3DAI%26time%3D" + str(
        timestamp) + "%26businessKey%3D" + tn + "%26trackingNumber%3D" + tn + "%26emailId%3Daleksei.shcherbakov%2540kuehne-nagel.com%26companyCode%3DCAKN"
    driver.get(url)
    delay = 20

    try:
        elem = WebDriverWait(driver, delay).until(
            EC.presence_of_element_located((By.XPATH,
                                            '/html/body/table[5]/tbody/tr/td[2]/form/table/tbody/tr/td[3]/table/tbody/tr[3]/td/table/tbody/tr[1]/td[2]/input')))

        elem = driver.find_element_by_xpath(
            '/html/body/table[5]/tbody/tr/td[2]/form/table/tbody/tr/td[3]/table/tbody/tr[3]/td/table/tbody/tr[1]/td[2]/input')
        elem.clear()
        elem.send_keys('aleksei.shcherbakov@can.win.int.kn')

        elem = driver.find_element_by_xpath(
            '/html/body/table[5]/tbody/tr/td[2]/form/table/tbody/tr/td[3]/table/tbody/tr[3]/td/table/tbody/tr[2]/td[2]/input')
        elem.clear()
        elem.send_keys('Password6')

        elem = driver.find_element_by_xpath(
            '/html/body/table[5]/tbody/tr/td[2]/form/table/tbody/tr/td[3]/table/tbody/tr[4]/td/input')
        elem.click()

        # driver.get(url)

    except TimeoutException:
        print("Loading took too much time!")

    try:
        elem = WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.XPATH,
                                            '/html/body/table[5]/tbody/tr/td[2]/form/table/tbody/tr[2]/td/table/tbody/tr[2]/td/table/tbody/tr[2]/td[4]/input')))
        status = driver.find_element_by_xpath(
            '/html/body/table[5]/tbody/tr/td[2]/form/table/tbody/tr[2]/td/table/tbody/tr[2]/td/table/tbody/tr[2]/td[4]/input').get_attribute(
            'value')
        if status == 'Unfiled':
            elem = driver.find_element_by_xpath('//*[@id="btn:submit:edit"]')
            elem.click()

        else:
            driver.close()
            return status

    except TimeoutException:
        print("Loading took too much time!")

    try:
        elem = WebDriverWait(driver, delay).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="fromPortcode"]')))
    except TimeoutException:
        print("Loading took too much time!")

    elem = driver.find_element_by_xpath('//*[@id="fromPortcode"]')
    elem.clear()
    elem.send_keys(from_port)

    elem = driver.find_element_by_xpath('//*[@id="toPortcode"]')
    elem.clear()
    elem.send_keys(to_port)

    elem = driver.find_element_by_xpath('//*[@id="goodsAt"]')
    elem.clear()
    elem.send_keys(goods_at)

    elem = driver.find_element_by_xpath('//*[@id="customsByLine1"]')
    if elem.get_attribute('value') == '':
        elem.clear()
        elem.send_keys('Kuehne + Nagel')

    elem = driver.find_element_by_xpath('//*[@id="cartageByLine1"]')
    if elem.get_attribute('value') == '':
        elem.clear()
        elem.send_keys('Kuehne + Nagel')

    select_obj = Select(driver.find_element_by_xpath('//*[@id="ctgTerms"]'))
    select_obj.select_by_visible_text('Pre Paid')

    elem = driver.find_element_by_xpath(
        '//*[@id="eta"]')
    elem.clear()
    elem.send_keys(eta.strftime('%d %b %Y'))

    elem = driver.find_element_by_xpath(
        '/html/body/table[5]/tbody/tr/td[2]/form/table[1]/tbody/tr/td[1]/table/tbody/tr[4]/td/table/tbody/tr[2]/td/table/tbody/tr[5]/td[5]/input')
    elem.clear()
    elem.send_keys('***PARS ETA WAREHOUSE ' + eta_warehouse.strftime('%d %b %Y %H:%M').upper() + '***')

    elem = driver.find_element_by_xpath(
        '//*[@id="otm"]')
    print(elem.get_attribute('checked'))
    if elem.get_attribute('checked') == 'false':
        elem.click()

    # change cargo control number if clear at the carrier
    if ccn_change:
        global mawb
        elem = driver.find_element_by_xpath('//*[@id="btn:submit:changeCcn"]')
        elem.click()

        try:
            elem = WebDriverWait(driver, delay).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="newCcn"]')))
        except TimeoutException:
            print("Loading took too much time!")

        elem = driver.find_element_by_xpath('//*[@id="newCcn"]')
        elem.send_keys(mawb)

        elem = driver.find_element_by_xpath('//*[@id="btn:submit:update"]')
        elem.click()

        try:
            elem = WebDriverWait(driver, delay).until(
                EC.presence_of_element_located((By.ID, 'btn:submit:save')))
        except TimeoutException:
            print("Loading took too much time!")

    elem = driver.find_element_by_id('btn:submit:save')
    elem.click()

    try:
        elem = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH,
                                            '/html/body/table[5]/tbody/tr/td[2]/form/table/tbody/tr[2]/td/table/tbody/tr[2]/td/table/tbody/tr[2]/td[4]/input')))
        status = driver.find_element_by_xpath(
            '/html/body/table[5]/tbody/tr/td[2]/form/table/tbody/tr[2]/td/table/tbody/tr[2]/td/table/tbody/tr[2]/td[4]/input').get_attribute(
            'value') + ' now'
        driver.close()


    except TimeoutException:
        print("Loading took too much time!")
        status = "Double check"

    return status


def cbsa_sublocation_checker(port, name):
    options = Options()
    options.add_argument("--headless")
    driver = webdriver.Firefox(firefox_options=options)
    driver.get('https://www.cbsa-asfc.gc.ca/import/codes/sw-ea-eng.html')

    delay = WebDriverWait(driver, 60).until(
        EC.presence_of_element_located(
            (By.CSS_SELECTOR, '#sufferance-table_filter > label:nth-child(1) > input:nth-child(1)')))

    elem = driver.find_element_by_css_selector('#sufferance-table_filter > label:nth-child(1) > input:nth-child(1)')
    elem.click()
    elem.send_keys('"' + name.lower() + '" ' + str(port))

    delay = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, '/html/body/main/div/table/tbody/tr/td[4]')))
    time.sleep(1)

    result = driver.find_element_by_xpath('/html/body/main/div/table/tbody/tr/td[5]').text
    driver.close()

    return result


def carrier_sublocation_checker(port, carrier, rule):
    goods_at = ''
    for line in range(13, 41):
        try:
            if carrier in rule['A' + str(line)].value:
                goods_at_name = rule['B' + str(line)].value
                try:
                    goods_at = cbsa_sublocation_checker(port, goods_at_name)
                    warning = ''
                    global ccn_change
                    ccn_change = True
                except:
                    warning = 'Can not get sublocation for %s, default used' % (goods_at_name)
                    goods_at = rule['E4'].value
                    goods_at_name = rule['E5'].value

        except TypeError:
            global ccn_change
            ccn_change = False
            goods_at = ''
            warning = ''
            goods_at_name = ''

    if goods_at == '':
        global ccn_change
        ccn_change = True
        warning = '%s sublocation missing, default used' % carrier
        goods_at = rule['E4'].value
        goods_at_name = rule['E5'].value

    return [goods_at, goods_at_name, warning]


def bso_sublocation_checker(port, bso, rule, carrier):
    goods_at = ''
    for line in range(2, 9):
        if rule['A' + str(line)].value == bso:

            goods_at_name = rule['B' + str(line)].value
            warning = 'Sublocation from BSO'
            if goods_at_name.lower() == 'carrier':
                return carrier_sublocation_checker(port, carrier, rule)
            try:
                goods_at = cbsa_sublocation_checker(port, goods_at_name)
            except:
                warning = ' Cannnot get sublocation for %s, default used; ' % (goods_at_name)
                goods_at = rule['E4'].value
                goods_at_name = rule['E5'].value
            return [goods_at, goods_at_name, warning]
    if goods_at == '':
        return False


def customer_sublocation_checker(port, customer, rule, carrier):
    goods_at = ''
    for line in range(13, 41):
        if consignee == rule['D' + str(line)].value:
            goods_at_name = rule['E' + str(line)].value
            if goods_at_name.lower() == 'carrier':
                return carrier_sublocation_checker(port, carrier, rule)
            try:
                goods_at = cbsa_sublocation_checker(port, goods_at_name)
                warning = 'Sublocation from Cnee'
            except:
                warning = 'Cannnot get sublocation for %s, default used' % (goods_at_name)
                goods_at = rule['E4'].value
                goods_at_name = rule['E5'].value
            return [goods_at, goods_at_name, warning]

    if goods_at == '':
        return False


if __name__ == '__main__':

    task_list = pd.read_csv('tasks.csv', error_bad_lines=False, encoding="ISO-8859-1", sep=';')
    task_list.to_excel('tasks.xlsx')
    rules = load_workbook('rules.xlsx', read_only=True)
    table = {}
    for index, task in task_list.iterrows():
        errors = []
        warnings = []
        goods_at = ''
        ccn_change = False

        try:
            tn = str(int(task['Tracking Number']))
            mawb = task['MAWB']
            mawb_link = '<a href="http://connect.track-trace.com/for/kn/aircargo/%s" target="_blank">%s</a>' % (
                mawb, mawb)
        except ValueError:
            continue

        team = task['Assigned Org. Team']
        branch = team[19:24]

        try:
            atd = datetime.strptime(str(task['Status 1300']), '%d-%b-%Y %H:%M')

        except ValueError:
            atd = 'Missing'
            warnings.append('Missing ATD')

        weight = task['Weight']
        destination = task['Destination']
        carrier = task['Carrier']
        bso = task['Product Group Code']
        consignee = task['Consignee code']
        try:
            rule = rules[destination]
        except ValueError:
            errors.append('Missing rules for %s airport' % destination)
            print('TN# %s contains following errors: %s' % (tn, errors))
            errors_string = ''
            for error in errors:
                errors_string += error + '; '
            try:
                table[
                    team] += '<tr><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td></tr>' % (
                    tn, mawb_link, task['HAWB'], task['Consignee code'], task['Milestone 1405'], destination, 'nan',
                    errors_string)
            except KeyError:
                table[
                    team] = '<tr><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td></tr>' % (
                    tn, mawb_link, task['HAWB'], task['Consignee code'], task['Milestone 1405'], destination,
                    'nan', errors_string)
            continue

        try:
            eta = datetime.strptime(str(task['Milestone 1405']), '%d-%b-%Y %H:%M')
            if eta.hour < int(rule['E7'].value):
                eta_warehouse = eta.replace(hour=int(rule['E8'].value), minute=0)
            else:
                eta_warehouse = eta.replace(day=eta.day + 1, hour=int(rule['E9'].value), minute=0)
        except ValueError:
            eta = 'Missing'
            errors.append('Missing ETA')

        from_port = rule['E1'].value
        to_port = rule['E1'].value

        customer_result = customer_sublocation_checker(from_port, consignee, rule, carrier)
        bso_result = bso_sublocation_checker(from_port, bso, rule, carrier)
        if customer_result:
            goods_at = customer_result[0]
            goods_at_name = customer_result[1]
            warnings.append(customer_result[2])
        elif bso_result:
            goods_at = bso_result[0]
            goods_at_name = bso_result[1]
            warnings.append(bso_result[2])
        elif int(weight) > int(rule['E3'].value):
            carrier_result = carrier_sublocation_checker(from_port, carrier, rule)
            goods_at = carrier_result[0]
            goods_at_name = carrier_result[1]
            warnings.append(carrier_result[2])
        else:
            goods_at = rule['E4'].value
            goods_at_name = rule['E5'].value

        if errors:
            print('TN# %s processed, status: %s; ' % (tn, errors))
            errors_string = ''
            for error in errors:
                errors_string += error + '; '
            try:
                table[
                    team] += '<tr><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td></tr>' % (
                    tn, mawb_link, task['HAWB'], task['Consignee code'], task['Milestone 1405'], destination, 'nan',
                    errors_string)
            except KeyError:
                table[
                    team] = '<tr><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td></tr>' % (
                    tn, mawb_link, task['HAWB'], task['Consignee code'], task['Milestone 1405'], destination,
                    'nan', errors_string)
        else:
            try:
                status = open_bond_page(tn, branch, from_port, to_port, goods_at, eta, eta_warehouse)
                warning_string = ''
                for warning in warnings:
                    warning_string += ', ' + warning
                errors = status + warning_string
                print('TN# %s contains following errors: %s' % (tn, errors))
                try:
                    table[
                        team] += '<tr><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td></tr>' % (
                        tn, mawb_link, task['HAWB'], task['Consignee code'], task['Milestone 1405'],
                        destination, goods_at_name + '(' + goods_at + ')', errors)
                except KeyError:
                    table[
                        team] = '<tr><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td></tr>' % (
                        tn, mawb_link, task['HAWB'], task['Consignee code'], task['Milestone 1405'],
                        destination,
                        goods_at_name + '(' + goods_at + ')', errors)
            except Exception as e:
                print('TN# %s finished with Error in processing: %s' % (tn, e))
                try:
                    table[
                        team] += '<tr><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td></tr>' % (
                        tn, mawb_link, task['HAWB'], task['Consignee code'], task['Milestone 1405'],
                        destination, goods_at_name + '(' + goods_at + ')', e)
                except KeyError:
                    table[
                        team] = '<tr><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td></tr>' % (
                        tn, mawb_link, task['HAWB'], task['Consignee code'], task['Milestone 1405'],
                        destination,
                        goods_at_name + '(' + goods_at + ')', e)

    for k, v in table.items():
        send_report(v, k)

    # open_bond_page('1021900126', 'CARBC', '0821', '0821', '4668', eta, eta_warehouse)
